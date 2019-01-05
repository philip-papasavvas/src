"""
Practice reading in Excel workbooks and choosing specific
cells and their values
"""

import openpyxl
import os

os.chdir('C:\\Users\\ppapasav\\Documents\\myPythonScripts')

workbook = openpyxl.load_workbook('data.xlsx')
data = workbook['data']

# data['A1']
# data['A1'].value # to access the value here

for i in range(1,10):
    print(i, data.cell(row=i, column = 5).value)

#create excel
wb = openpyxl.Workbook()
#sheet  = wb.sheetnames # get_sheet_names deprecated
sheet = wb['Sheet']
sheet['A1'].value = 32
import os
os.chdir("C:\\Users\\ppapasav\\Documents\\myPythonScripts")
wb.save('example.xlsx')

sheet2 = wb.create_sheet()
wb.sheetnames
sheet2.title = "newSheetName"
wb.save('example_one.xlsx')
wb.create_sheet(index=0, title = 'myOtherSheet')
wb.save('example3.xlsx')